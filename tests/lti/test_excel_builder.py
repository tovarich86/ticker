import io
import pytest
import pandas as pd
import openpyxl
from datetime import date, datetime
from src.lti.config import OUTORGAS
from src.lti.engine import ApuracaoResult, TickerResult, _calcular_grupos
from src.lti.excel_builder import gerar_excel_bytes


def _make_resultado(ano: int = 2023) -> ApuracaoResult:
    cfg = OUTORGAS[ano]
    tickers_data = [
        TickerResult(
            ticker="VALE3", ticker_original="VALE3",
            vwap_p0=50.0, vwap_pf=60.0, dividendos_total=5.0,
            tsr=0.30, rank=1, grupo=1, status="INCLUIDO",
            df_cotacoes_p0=pd.DataFrame([{"Ticker": "VALE3", "Date": date(2023, 3, 2), "Average": 50.0, "Quantity": 1000}]),
            df_cotacoes_pf=pd.DataFrame([{"Ticker": "VALE3", "Date": date(2026, 3, 2), "Average": 60.0, "Quantity": 1000}]),
            df_dividendos=pd.DataFrame([{"Ticker": "VALE3", "lastDatePriorEx": "15/06/2023", "value": 5.0}]),
            df_bonificacoes=pd.DataFrame(),
            divs_ajustados=[{"Data Ex": "15/06/2023", "Valor/Ação (R$)": 5.0, "Multiplicador": 1.0, "Total Recebido (R$)": 5.0}],
        ),
        TickerResult(
            ticker="AZUL4", ticker_original="AZUL4",
            vwap_p0=None, vwap_pf=None, dividendos_total=0.0,
            status="EXCLUIDO_FORCADO", motivo_exclusao="exclusao_forcada em config",
        ),
    ]
    incluidos = [t for t in tickers_data if t.status == "INCLUIDO"]
    grupos = _calcular_grupos(incluidos)

    return ApuracaoResult(
        outorga=cfg,
        tickers=tickers_data,
        ranking=incluidos,
        grupos=grupos,
        n_incluidos=1,
        n_excluidos=1,
        timestamp=datetime(2026, 4, 10, 12, 0),
    )


def _abas(xlsx_bytes: bytes) -> list[str]:
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))
    return wb.sheetnames


def test_gerar_excel_contem_todas_abas():
    resultado = _make_resultado()
    xlsx = gerar_excel_bytes(resultado)
    abas = _abas(xlsx)
    for esperada in ["Resultado", "Grupos", "Composição", "Metodologia",
                     "Cotacao_P0", "Cotacao_Pf", "Dividendos", "Eventos_Corp",
                     "DivAjustados", "Exclusoes", "Divergencias_YF", "Config"]:
        assert esperada in abas, f"Aba '{esperada}' não encontrada"


def test_resultado_aba_tem_ticker_incluido():
    resultado = _make_resultado()
    xlsx = gerar_excel_bytes(resultado)
    wb = openpyxl.load_workbook(io.BytesIO(xlsx))
    ws = wb["Resultado"]
    tickers_na_aba = [ws.cell(row=r, column=2).value for r in range(2, ws.max_row + 1)]
    assert "VALE3" in tickers_na_aba


def test_exclusoes_aba_tem_excluidos():
    resultado = _make_resultado()
    xlsx = gerar_excel_bytes(resultado)
    wb = openpyxl.load_workbook(io.BytesIO(xlsx))
    ws = wb["Exclusoes"]
    tickers_excluidos = [ws.cell(row=r, column=1).value for r in range(2, ws.max_row + 1)]
    assert "AZUL4" in tickers_excluidos


def test_config_aba_tem_ano_outorga():
    resultado = _make_resultado(2023)
    xlsx = gerar_excel_bytes(resultado)
    wb = openpyxl.load_workbook(io.BytesIO(xlsx))
    ws = wb["Config"]
    all_values = [ws.cell(row=r, column=2).value for r in range(1, ws.max_row + 1)]
    assert 2023 in all_values
